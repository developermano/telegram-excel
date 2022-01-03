import os
from telegram import Update,Bot
from telegram.ext import Updater, CommandHandler, CallbackContext
from replit import db
import openpyxl
from server import keepalive
keepalive()

bot=Bot(os.environ['token'])

def start(update: Update, context: CallbackContext) -> None:
    update.message.reply_text('hello ! \ncommands: \n  /setusername [username] -- setup the account. \n /expirydate --  get expiry date')

def setup(update: Update, context: CallbackContext) -> None:
    if len(context.args)!=0:
      db[str(update.effective_message.chat_id)] = context.args[0]
      update.message.reply_text('your account is added successfully .')
def expirydate(update: Update, context: CallbackContext) -> None:

  if db[str(update.effective_message.chat_id)]=="":
    update.message.reply_text('sorry . please run /setusername to set your username')

  username=db[str(update.effective_message.chat_id)]

  wb_obj = openpyxl.load_workbook("example.xlsx")
  sheet_obj = wb_obj.active
  max_col = sheet_obj.max_column
  max_row = sheet_obj.max_row
 
# Loop will print all columns name
  for i in range(1, max_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    print(cell_obj.value)
    if str(cell_obj.value)==str(username):
      update.message.reply_text("your expiry date is "+str(sheet_obj.cell(row = i, column = 3).value))
      return
  update.message.reply_text('i could not find your username. please give correct username.')

def main() -> None:
    """Run the bot."""
    # Create the Updater and pass it your bot's token.

    
    my_secret = os.environ['token']
    updater = Updater(my_secret)

    # Get the dispatcher to register handlers
    dispatcher = updater.dispatcher
    dispatcher.add_handler(CommandHandler("start", start))
    dispatcher.add_handler(CommandHandler("setusername", setup))
    dispatcher.add_handler(CommandHandler("expirydate", expirydate))
    updater.start_polling()
    updater.idle()


if __name__ == '__main__':
    main()